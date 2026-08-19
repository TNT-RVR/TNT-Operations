#!/usr/bin/env python3
"""
Turn the "Checklist" spreadsheet into SQL for `field_checklist`.

The sync (netlify/functions/checklist-sheet-sync.mjs) would import the same rows
on its first run, once a service account exists and the sheet is shared with it.
This exists so the app can hold the real marks TODAY, and so a season nobody
syncs any more — 2023, 2024 — is loaded once from the file rather than requiring
the sheet to stay reachable forever.

    python scripts/import_checklist.py "C:/Users/tyler/Downloads/Checklist (1).xlsx"
    # then paste scripts/checklist_import.sql into the Supabase SQL editor

Reading the FILL matters as much as the value. TNT marks a step done by
highlighting the cell blue; the date alone cannot say whether it is a plan or a
record, which is the whole reason the app splits them into two columns.

Output is idempotent: `on conflict (year, field_name, step) do update`, so
re-running after fixing a cell corrects rather than duplicates. It deliberately
leaves `synced_*` NULL — these rows have never agreed with the sheet through the
sync, and claiming they had would make the first real sync treat a later
spreadsheet edit as "no change" and drop it.
"""
from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is needed: python -m pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "scripts" / "checklist_import.sql"

# Sheet header → the step key the app stores. Anything else (Gallons,
# Structures, Image, Type, Blocks) is not a step and is left alone.
STEPS = {
    "flag": "flag",
    "structures in": "structures_in",
    "mouse poison": "mouse_poison",
    "bees in": "bees_in",
    "structures out": "structures_out",
}


def is_done_fill(cell) -> bool:
    """
    Blue fill = completed. Tolerant on purpose: three seasons of hand
    highlighting are not one exact blue, and openpyxl reports themed colours as
    objects rather than RGB strings.
    """
    fill = cell.fill
    if not fill or fill.patternType != "solid":
        return False
    rgb = getattr(fill.start_color, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) < 6:
        return False
    r, g, b = int(rgb[-6:-4], 16), int(rgb[-4:-2], 16), int(rgb[-2:], 16)
    if r > 230 and g > 230 and b > 230:  # white
        return False
    return b > 128 and b > r + 40 and b > g + 25


def as_iso(value) -> str | None:
    """A real date, or None. Text like 'Half- 7/16/2026' is a note, not a date."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if m:
        mm, dd, yy = m.groups()
        return f"{yy}-{int(mm):02d}-{int(dd):02d}"
    return None


def sql_str(s: str) -> str:
    return "'" + str(s).replace("'", "''") + "'"


def sql_date(d: str | None) -> str:
    return f"date {sql_str(d)}" if d else "null"


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    book = Path(sys.argv[1])
    if not book.exists():
        sys.exit(f"no such file: {book}")

    wb = openpyxl.load_workbook(book)
    rows_sql: list[str] = []
    summary: list[str] = []

    for ws in wb.worksheets:
        year = ws.title.strip()
        if not re.fullmatch(r"\d{4}", year):
            print(f"  skipping sheet {ws.title!r} — not a season")
            continue

        header = [str(c.value or "").strip().lower() for c in ws[1]]
        cols = {STEPS[h]: i for i, h in enumerate(header) if h in STEPS}
        try:
            name_col = header.index("field name")
        except ValueError:
            name_col = 0
        if not cols:
            print(f"  skipping {year} — no step columns found")
            continue

        planned = done = notes = 0
        for row in ws.iter_rows(min_row=2):
            if name_col >= len(row):
                continue
            field_name = str(row[name_col].value or "").strip()
            if not field_name:
                continue
            for step, ci in cols.items():
                if ci >= len(row):
                    continue
                cell = row[ci]
                if cell.value in (None, ""):
                    continue
                iso = as_iso(cell.value)
                note = "" if iso else str(cell.value).strip()
                completed = iso if (iso and is_done_fill(cell)) else None
                plan = None if completed else iso
                if not (completed or plan or note):
                    continue
                planned += 1 if plan else 0
                done += 1 if completed else 0
                notes += 1 if note else 0
                rows_sql.append(
                    f"  ({sql_str(year)}, {sql_str(field_name)}, {sql_str(step)}, "
                    f"{sql_date(plan)}, {sql_date(completed)}, {sql_str(note)})"
                )
        summary.append(f"  {year}: {done} done, {planned} planned, {notes} with a note")

    if not rows_sql:
        sys.exit("nothing to import — no step columns matched")

    body = ",\n".join(rows_sql)
    OUT.write_text(
        "-- Generated by scripts/import_checklist.py — paste into the Supabase SQL editor.\n"
        f"-- Source: {book.name}\n"
        "--\n"
        "-- Idempotent: re-running corrects rows rather than duplicating them. The\n"
        "-- synced_* columns are left NULL on purpose — these marks have never been\n"
        "-- through the sync, and pretending otherwise would make the first real sync\n"
        "-- read a later spreadsheet edit as 'nothing changed' and discard it.\n\n"
        "insert into public.field_checklist\n"
        "  (year, field_name, step, planned_date, completed_date, note)\n"
        "values\n" + body + "\n"
        "on conflict (year, field_name, step) do update set\n"
        "  planned_date   = excluded.planned_date,\n"
        "  completed_date = excluded.completed_date,\n"
        "  note           = excluded.note;\n",
        encoding="utf-8",
    )
    print(f"wrote {OUT.relative_to(ROOT)} — {len(rows_sql)} marks")
    print("\n".join(summary))


if __name__ == "__main__":
    main()
